from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path #libreria para saber si un archivo/ubicacion existe
import tkinter as tk
from tkinter import messagebox

#crear app grafica pizzeria:
#-3 gustos: carne, jamon y queso, pollo
#tomar los pedidos y almacenarlos en un archivo txt/excel. Un txt por cada pedido que diga:
#cuantas empanadas de cada gusto pidio

######  FUNCIONES: ######
                
def calcularPrecio(cantidad):
    if cantidad>=12:
        resto = cantidad%12
        docena = (cantidad-resto)/12
        return resto*40+docena*400
    else:
        return cantidad * 40

def guardarDatos(pedido):
    wb.active.append(pedido)
    wb.save("empanadas.xlsx")
    

def convertirDatos(dato):
    try:
        dato = int(dato)
    except (TypeError,ValueError):
        messagebox.showerror("Error de cantidad","Debe ingresar números enteros en la cantidad de empanadas.")
        dato = "error"
        borrarEntrada()
    return dato

def comprobarArchivo():
    existe = Path("empanadas.xlsx") #Busco el documento
    if existe.is_file(): #Si el documento existe... lo cargo y uso
        wb = load_workbook(filename = "empanadas.xlsx")
        ws = wb.active
    else: #Si no existe... Lo creo
        wb = Workbook()
        ws = wb.active
        titulo = ("Nombre","Carne","JyQ","Pollo","Precio Total")
        ws.append(titulo)
    return wb 
    
def realizarPedido():
    total = 0
    if not txtNombre.get():
        txtNombre.focus_set()
        messagebox.showwarning("Faltan Datos", "¡Debe ingresar su nombre para completar el pedido!")
    else:
        pedidoCantidad = [txtCarne.get(),txtPollo.get(),txtJq.get()]
        for p in pedidoCantidad:
            if(convertirDatos(p)=="error"):
                total = 0
                break
            else:
                p = convertirDatos(p)
                total = total + p
        total = calcularPrecio(total)  
        if total==0:
            borrarEntrada()
        elif messagebox.askyesno("Confirmar pedido","El total de su pedido es $"+str(total)+" ¿Desea confirmar?"):   
            pedidoCantidad.insert(0,txtNombre.get())
            pedidoCantidad.append(total)
            guardarDatos(pedidoCantidad)
            messagebox.showinfo("Éxito","Pedido Realizado")
            borrarEntrada()
        else:
            borrarEntrada()
            messagebox.showinfo("Cancelar","Pedido cancelado.")

def borrarEntrada():
    txtCarne.delete(0,tk.END)
    txtJq.delete(0,tk.END)
    txtPollo.delete(0,tk.END)
    txtNombre.delete(0,len(txtNombre.get()))
    txtCarne.insert(0,"0")
    txtJq.insert(0,"0")
    txtPollo.insert(0,"0")
    
    
    

##### MAIN #####


v = tk.Tk()
v.config(height=300, width=500)  
v.title("Empanadas - by Cande")
txtCarne = tk.Entry()
txtCarne.place(x=300,y=50)
txtCarne.insert(tk.END,"0")
txtPollo = tk.Entry()
txtPollo.place(x=300,y=100)
txtPollo.insert(tk.END,"0")
txtJq = tk.Entry()
txtJq.place(x=300,y=150)
txtJq.insert(tk.END,"0")
txtNombre = tk.Entry()
txtNombre.place(x=300,y=200)
aceptar = tk.Button(text="Aceptar",command=realizarPedido)
cancelar = tk.Button(text="Cancelar", command=borrarEntrada)
aceptar.place(x=100,y=250)
cancelar.place(x=300,y=250)
tk.Label(text=" ¡ Bienvenidx ! ").place(x=200,y=25)
lblCarne = tk.Label(text="Ingrese cantidad de carne: ")
lblCarne.place(x=100,y=50)
lblPollo = tk.Label(text="Ingrese cantidad de pollo: ")
lblPollo.place(x=100,y=100)
lblJq = tk.Label(text="Ingrese cantidad de jamón y queso: ")
lblJq.place(x=100,y=150)
lblNombre = tk.Label(text="Ingrese su nombre: ").place(x=100,y=200)
wb = comprobarArchivo()
v.mainloop()





