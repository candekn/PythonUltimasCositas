from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path #libreria para saber si un archivo/ubicacion existe

"""
crear app grafica pizzeria:
-3 gustos: carne, jamon y queso, pollo
tomar los pedidos y almacenarlos en un archivo txt/excel. Un txt por cada pedido que diga:
cuantas empanadas de cada gusto pidio
"""
######  FUNCIONES: ######
def pedidoCantidad(sabor):
    cant = input("Ha seleccionado '"+sabor +"' ¿Cuántas va a llevar? (presione 0 para cancelar): ")
    cant = int(cant)
    return cant
                
def calcularPrecio(cantidad):
    if cantidad>=12:
        resto = cantidad%12
        docena = (cantidad-resto)/12
        return resto*40+docena*400
    else:
        return cantidad * 40

##### MAIN #####

existe = Path("empanadas.xlsx") #Busco el documento
if existe.is_file(): #Si el documento existe... lo cargo y uso
    wb = load_workbook(filename = "empanadas.xlsx")
    ws = wb.active
else: #Si no existe... Lo creo
    wb = Workbook()
    ws = wb.active
    titulo = ("Nombre","Carne","JyQ","Pollo","Precio Total")
    ws.append(titulo)
###Defino las variables ####
cant = 0
carne = 0 
jq = 0
pollo = 0
nombre = ""
### Empieza la ejecucion ####
while True:
    print("Bienvenido al pedido de empanadas. Cada empanada sale $40, y la docena $400")
    nombre = input("¿Cuál es su nombre? (marque x para cancelar): ")
    if nombre.lower()=='x':
        print("Gracias por utilizar el servicio de pedidos Cande :D \n\n")
        break
    while True:
        print("Elija el sabor que desee:\n-C: Carne\n-JQ: Jamón y Queso\n-P: Pollo")
        sabor=input("Escriba la opcion deseada (C/JQ/P o N para terminar): ")
        sabor = sabor.upper()
        if sabor=="N":
            total = 0
            total = total + (calcularPrecio(carne)) + calcularPrecio(pollo) + calcularPrecio(jq)
            print("El precio total es $",total)
            aceptarP = input("¿Desea confirmar el pedido? Y/N: ")
            if(aceptarP.upper()=="Y"):
                persona = [nombre,carne,jq,pollo,total]
                ws.append(persona)
                wb.save("empanadas.xlsx")
                print("Pedido realizado!\n\n")
                break
            else:
                print("Pedido Cancelado\n")
                break    
        elif sabor!="C" and sabor!="JQ" and sabor!="P":
            print("El sabor elegido no es válido")
        else:
            if sabor=="C":
                carne = carne + pedidoCantidad(sabor)
            elif sabor=="JQ":
                jq = jq + pedidoCantidad(sabor)
            elif sabor =="P":
                pollo = pollo + pedidoCantidad(sabor)
        
